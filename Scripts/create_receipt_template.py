"""
請書Excelテンプレート作成スクリプト
Power Automateで差し込み可能なExcelテンプレートを生成

プレースホルダー:
- {{VendorName}}: 業者名（B3）
- {{OrderDate}}: 発注日（G5）
- {{ItemName}}: 品目（B10）
- {{Manufacturer}}: メーカー（C10）
- {{Quantity}}: 数量（D10）
- {{EstimatedAmount}}: 見積額（F10）

計算セル:
- E10: 単価（=F10/D10）
- F31: 小計（=SUM(F10:F30)）
- F32: 消費税（=F31*0.1）
- F33: 合計（=F31+F32）

業者入力セル（編集可能）:
- G10:G20: 納期・備考欄
- B35: 担当者サイン
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
import os

def create_receipt_template():
    """請書Excelテンプレートを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = '発注請書'
    
    # スタイル定義
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    title_font = Font(size=18, bold=True)
    
    # 列幅設定
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    
    # タイトル
    ws['D1'] = '発 注 請 書'
    ws['D1'].font = title_font
    ws['D1'].alignment = Alignment(horizontal='center')
    
    # 宛先（自社）
    ws['B3'] = 'セルジェンテック株式会社　御中'
    ws['B3'].font = Font(size=12)
    
    # 発注日
    ws['F5'] = '発注日：'
    ws['G5'] = '{{OrderDate}}'
    ws['G5'].alignment = Alignment(horizontal='left')
    
    # 業者名（差し込み）
    ws['F3'] = '貴社名：'
    ws['G3'] = '{{VendorName}}'
    
    # 本文
    ws['B7'] = '下記の通り発注を受領いたしましたことをご報告申し上げます。'
    
    # 明細ヘッダー
    headers = ['No.', '品目', 'メーカー', '数量', '単価', '金額', '納期・備考']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # 明細行（1行目：差し込みデータ）
    ws['A10'] = 1
    ws['A10'].border = thin_border
    ws['A10'].alignment = Alignment(horizontal='center')
    
    # 品目
    ws['B10'] = '{{ItemName}}'
    ws['B10'].border = thin_border
    
    # メーカー
    ws['C10'] = '{{Manufacturer}}'
    ws['C10'].border = thin_border
    
    # 数量
    ws['D10'] = '{{Quantity}}'
    ws['D10'].border = thin_border
    ws['D10'].alignment = Alignment(horizontal='right')
    
    # 単価（計算式: 金額÷数量）- 差し込み後に計算
    ws['E10'] = '=IF(D10<>"",F10/D10,"")'
    ws['E10'].border = thin_border
    ws['E10'].alignment = Alignment(horizontal='right')
    ws['E10'].number_format = '#,##0'
    
    # 金額（見積額）
    ws['F10'] = '{{EstimatedAmount}}'
    ws['F10'].border = thin_border
    ws['F10'].alignment = Alignment(horizontal='right')
    
    # 納期・備考（業者入力欄）
    ws['G10'] = ''
    ws['G10'].border = thin_border
    ws['G10'].protection = Protection(locked=False)  # 編集可能
    
    # 明細行（2～21行：空行）
    for row in range(11, 31):
        ws.cell(row=row, column=1).value = row - 9 if row < 21 else ''
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 7:  # 納期・備考欄は編集可能
                cell.protection = Protection(locked=False)
    
    # 小計
    ws['E31'] = '小計'
    ws['E31'].font = header_font
    ws['E31'].alignment = Alignment(horizontal='right')
    ws['F31'] = '=SUM(F10:F30)'
    ws['F31'].border = thin_border
    ws['F31'].number_format = '#,##0'
    
    # 消費税
    ws['E32'] = '消費税（10%）'
    ws['E32'].font = header_font
    ws['E32'].alignment = Alignment(horizontal='right')
    ws['F32'] = '=F31*0.1'
    ws['F32'].border = thin_border
    ws['F32'].number_format = '#,##0'
    
    # 合計
    ws['E33'] = '合計'
    ws['E33'].font = Font(bold=True, size=12)
    ws['E33'].alignment = Alignment(horizontal='right')
    ws['F33'] = '=F31+F32'
    ws['F33'].border = thin_border
    ws['F33'].font = Font(bold=True, size=12)
    ws['F33'].number_format = '#,##0'
    
    # 担当者サイン欄
    ws['B35'] = '受領担当者：'
    ws['C35'] = ''
    ws['C35'].border = Border(bottom=Side(style='thin'))
    ws['C35'].protection = Protection(locked=False)  # 編集可能
    ws.merge_cells('C35:E35')
    
    # 日付欄
    ws['F35'] = '日付：'
    ws['G35'] = ''
    ws['G35'].border = Border(bottom=Side(style='thin'))
    ws['G35'].protection = Protection(locked=False)  # 編集可能
    
    # シート保護設定（業者入力欄以外をロック）- パスワードなし
    ws.protection.sheet = True
    
    return wb

if __name__ == '__main__':
    output_path = r'c:\Users\千賀聡志\OneDrive - セルジェンテック株式会社\sharepointリスト化\Documents\02_Templates\請書テンプレート.xlsx'
    
    wb = create_receipt_template()
    wb.save(output_path)
    
    print(f'✅ 請書Excelテンプレートを作成しました: {output_path}')
    print()
    print('プレースホルダー一覧:')
    print('  - {{VendorName}}: 業者名（G3）')
    print('  - {{OrderDate}}: 発注日（G5）')
    print('  - {{ItemName}}: 品目（B10）')
    print('  - {{Manufacturer}}: メーカー（C10）')
    print('  - {{Quantity}}: 数量（D10）')
    print('  - {{EstimatedAmount}}: 見積額（F10）')
    print()
    print('計算セル:')
    print('  - E10: 単価（=F10/D10）')
    print('  - F31: 小計')
    print('  - F32: 消費税（10%）')
    print('  - F33: 合計')
    print()
    print('業者入力セル（編集可能）:')
    print('  - G10:G30: 納期・備考欄')
    print('  - C35:E35: 担当者サイン')
    print('  - G35: 日付')
