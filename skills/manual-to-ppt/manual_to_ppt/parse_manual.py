"""Parse Excel manuals and extract text/image content."""

from __future__ import annotations

import io
import os
from pathlib import Path
from typing import Any

import openpyxl
from PIL import Image as PILImage


class ExcelManualParser:
    """Parse Excel manuals into structured text/image content."""

    def __init__(self, excel_path: str, temp_dir: str = "temp/images") -> None:
        self.excel_path = excel_path
        self.temp_dir = temp_dir
        self.workbook = None
        self.worksheet = None
        self.content_data: list[dict[str, Any]] = []
        self.images: list[dict[str, Any]] = []

    def load_workbook(self) -> bool:
        """Load workbook and active worksheet."""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.worksheet = self.workbook.active
            return True
        except Exception as exc:  # noqa: BLE001
            print(f"ERROR: failed to load workbook: {exc}")
            return False

    def extract_text_content(self) -> None:
        """Extract non-empty cell content."""
        assert self.worksheet is not None
        for row_idx, row in enumerate(self.worksheet.iter_rows(values_only=False), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is None:
                    continue
                value = str(cell.value).strip()
                if not value:
                    continue
                self.content_data.append(
                    {
                        "row": row_idx,
                        "col": col_idx,
                        "value": value,
                        "font_size": cell.font.size if cell.font else None,
                        "is_bold": bool(cell.font.bold) if cell.font else False,
                        "type": "text",
                    }
                )

    def extract_images(self) -> None:
        """Extract embedded worksheet images to temporary files."""
        assert self.worksheet is not None
        os.makedirs(self.temp_dir, exist_ok=True)
        image_count = 0
        for img in getattr(self.worksheet, "_images", []):
            try:
                image_count += 1
                image_data = img._data()
                pil_image = PILImage.open(io.BytesIO(image_data))
                anchor = img.anchor
                row = anchor._from.row if hasattr(anchor, "_from") else 0
                col = anchor._from.col if hasattr(anchor, "_from") else 0

                image_filename = f"image_{image_count}_row{row}_col{col}.png"
                image_path = os.path.join(self.temp_dir, image_filename)
                pil_image.save(image_path, "PNG")

                self.images.append(
                    {
                        "filename": image_filename,
                        "path": image_path,
                        "row": row,
                        "col": col,
                        "width": pil_image.width,
                        "height": pil_image.height,
                        "type": "image",
                    }
                )
            except Exception as exc:  # noqa: BLE001
                print(f"WARN: image extraction failed: {exc}")

    def organize_content(self) -> list[dict[str, Any]]:
        """Return merged content sorted by row and column."""
        all_content = self.content_data + self.images
        all_content.sort(key=lambda item: (item["row"], item["col"]))
        return all_content

    def parse(self) -> list[dict[str, Any]] | bool:
        """Run full parse pipeline."""
        if not self.load_workbook():
            return False
        self.extract_text_content()
        self.extract_images()
        return self.organize_content()


def main() -> int:
    """Legacy standalone parser entrypoint."""
    module_path = Path(__file__).resolve()
    repo_root = module_path.parents[3]
    excel_path = repo_root / "05_mail" / "相見積操作マニュアル.xlsx"
    temp_dir = repo_root / "05_mail" / "temp" / "images"
    if not excel_path.exists():
        print(f"ERROR: file not found: {excel_path}")
        return 1

    parser = ExcelManualParser(str(excel_path), str(temp_dir))
    result = parser.parse()
    return 0 if result else 1


if __name__ == "__main__":
    raise SystemExit(main())

