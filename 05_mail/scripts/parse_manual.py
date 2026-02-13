"""
Excelマニュアル解析スクリプト

このスクリプトは、Excelファイルから以下の情報を抽出します：
- シート構成
- テキストコンテンツ（セル内容、見出し、手順）
- 埋め込み画像（スクリーンショット）
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

# Windows環境でのUnicodeエンコーディング設定
if sys.platform == 'win32':
    import codecs
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')


class ExcelManualParser:
    """Excelマニュアルを解析するクラス"""

    def __init__(self, excel_path, temp_dir="temp/images"):
        """
        初期化

        Args:
            excel_path: Excelファイルのパス
            temp_dir: 画像を一時保存するディレクトリ
        """
        self.excel_path = excel_path
        self.temp_dir = temp_dir
        self.workbook = None
        self.worksheet = None
        self.content_data = []
        self.images = []

    def load_workbook(self):
        """Excelワークブックを読み込む"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            print(f"✓ Excelファイルを読み込みました: {self.excel_path}")
            print(f"  シート名: {self.workbook.sheetnames}")

            # アクティブシートを使用
            self.worksheet = self.workbook.active
            print(f"  使用するシート: {self.worksheet.title}")
            print(f"  行数: {self.worksheet.max_row}, 列数: {self.worksheet.max_column}")
            return True
        except Exception as e:
            print(f"✗ エラー: Excelファイルの読み込みに失敗しました: {e}")
            return False

    def extract_text_content(self):
        """テキストコンテンツを抽出"""
        print("\n--- テキストコンテンツの抽出 ---")

        for row_idx, row in enumerate(self.worksheet.iter_rows(values_only=False), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None and str(cell.value).strip():
                    cell_data = {
                        'row': row_idx,
                        'col': col_idx,
                        'value': str(cell.value).strip(),
                        'font_size': cell.font.size if cell.font else None,
                        'is_bold': cell.font.bold if cell.font else False,
                        'type': 'text'
                    }
                    self.content_data.append(cell_data)

        print(f"✓ テキストセル数: {len(self.content_data)}")

        # サンプル表示（最初の10件）
        print("\n  サンプルテキスト（最初の10件）:")
        for i, data in enumerate(self.content_data[:10], 1):
            value_preview = data['value'][:50] + '...' if len(data['value']) > 50 else data['value']
            print(f"    {i}. Row {data['row']}, Col {data['col']}: {value_preview}")

    def extract_images(self):
        """埋め込み画像を抽出"""
        print("\n--- 画像の抽出 ---")

        # 一時ディレクトリを作成
        os.makedirs(self.temp_dir, exist_ok=True)
        print(f"  画像保存先: {self.temp_dir}")

        # ワークシートから画像を抽出
        if hasattr(self.worksheet, '_images'):
            image_count = 0
            for img in self.worksheet._images:
                try:
                    image_count += 1
                    # 画像データを取得
                    image_data = img._data()

                    # PILImageとして読み込み
                    pil_image = PILImage.open(io.BytesIO(image_data))

                    # 画像のアンカー位置を取得
                    anchor = img.anchor
                    row = anchor._from.row if hasattr(anchor, '_from') else 0
                    col = anchor._from.col if hasattr(anchor, '_from') else 0

                    # 画像を保存
                    image_filename = f"image_{image_count}_row{row}_col{col}.png"
                    image_path = os.path.join(self.temp_dir, image_filename)
                    pil_image.save(image_path, 'PNG')

                    image_info = {
                        'filename': image_filename,
                        'path': image_path,
                        'row': row,
                        'col': col,
                        'width': pil_image.width,
                        'height': pil_image.height,
                        'type': 'image'
                    }
                    self.images.append(image_info)

                    print(f"    {image_count}. {image_filename} ({pil_image.width}x{pil_image.height}) at Row {row}, Col {col}")

                except Exception as e:
                    print(f"    ✗ 画像 {image_count} の抽出に失敗: {e}")

        print(f"✓ 抽出した画像数: {len(self.images)}")

    def organize_content(self):
        """コンテンツを整理して構造化"""
        print("\n--- コンテンツの整理 ---")

        # テキストと画像を行番号でソート
        all_content = self.content_data + self.images
        all_content.sort(key=lambda x: (x['row'], x['col']))

        print(f"✓ 総コンテンツ数: {len(all_content)}")
        print(f"  テキスト: {len(self.content_data)}, 画像: {len(self.images)}")

        return all_content

    def get_summary(self):
        """解析結果のサマリーを取得"""
        return {
            'excel_path': self.excel_path,
            'sheet_name': self.worksheet.title if self.worksheet else None,
            'max_row': self.worksheet.max_row if self.worksheet else 0,
            'max_col': self.worksheet.max_column if self.worksheet else 0,
            'text_count': len(self.content_data),
            'image_count': len(self.images),
            'temp_dir': self.temp_dir
        }

    def parse(self):
        """全体の解析を実行"""
        print("=" * 60)
        print("Excelマニュアル解析開始")
        print("=" * 60)

        if not self.load_workbook():
            return False

        self.extract_text_content()
        self.extract_images()
        organized_content = self.organize_content()

        print("\n" + "=" * 60)
        print("解析完了")
        print("=" * 60)

        summary = self.get_summary()
        print(f"\nサマリー:")
        print(f"  ファイル: {summary['excel_path']}")
        print(f"  シート: {summary['sheet_name']}")
        print(f"  サイズ: {summary['max_row']}行 x {summary['max_col']}列")
        print(f"  テキスト: {summary['text_count']}件")
        print(f"  画像: {summary['image_count']}件")
        print(f"  画像保存先: {summary['temp_dir']}")

        return organized_content


def main():
    """メイン処理"""
    # Excelファイルのパス
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    excel_path = project_dir / "相見積操作マニュアル.xlsx"
    temp_dir = project_dir / "temp" / "images"

    print(f"対象ファイル: {excel_path}")
    print(f"画像保存先: {temp_dir}\n")

    if not excel_path.exists():
        print(f"✗ エラー: ファイルが見つかりません: {excel_path}")
        return 1

    # 解析実行
    parser = ExcelManualParser(str(excel_path), str(temp_dir))
    content = parser.parse()

    if content:
        print(f"\n✓ 解析が正常に完了しました")
        return 0
    else:
        print(f"\n✗ 解析に失敗しました")
        return 1


if __name__ == "__main__":
    sys.exit(main())
